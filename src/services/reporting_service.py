from __future__ import annotations

import logging
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill

from src.core.config import AppConfig
from src.core.constants import DEFAULT_COLUMN_WIDTH_SAMPLE_ROWS, DEFAULT_MAX_COLUMN_WIDTH
from src.core.exceptions import ReportingError
from src.models.pipeline_metrics import PipelineMetrics
from src.models.report_models import ReportArtifact

LOGGER = logging.getLogger("etl.reporting")


class ReportingService:
    def __init__(self, config: AppConfig) -> None:
        self.config = config

    def generate_report(
        self,
        *,
        metrics: PipelineMetrics,
        analysis_results: dict[str, Any],
        cleaned_df: pd.DataFrame | None,
        cleaned_data_path: Path | None = None,
    ) -> ReportArtifact:
        report_path = self._resolve_report_path(metrics.run_id)
        exported_cleaned_data_path = cleaned_data_path
        drop_reasons_path = self._resolve_drop_reasons_path(metrics.run_id)
        summary_path = self._resolve_summary_path(metrics.run_id)

        try:
            self._write_drop_reasons(metrics, drop_reasons_path)
            self._write_structured_summary(
                metrics=metrics,
                analysis_results=analysis_results,
                report_path=report_path,
                cleaned_data_path=exported_cleaned_data_path,
                drop_reasons_path=drop_reasons_path,
                summary_path=summary_path,
            )

            if self.config.report_format == "csv":
                summary = self._build_pipeline_summary_frame(metrics)
                summary.to_csv(report_path, index=False)
                LOGGER.info("CSV report generated.", extra={"stage": "reporting", "report_path": str(report_path)})
                return ReportArtifact(
                    report_path=report_path,
                    cleaned_data_path=exported_cleaned_data_path,
                    drop_reasons_path=drop_reasons_path,
                    summary_path=summary_path,
                )

            if self.config.report_format == "parquet":
                summary = self._build_pipeline_summary_frame(metrics)
                summary.to_parquet(report_path, index=False)
                LOGGER.info("Parquet report generated.", extra={"stage": "reporting", "report_path": str(report_path)})
                return ReportArtifact(
                    report_path=report_path,
                    cleaned_data_path=exported_cleaned_data_path,
                    drop_reasons_path=drop_reasons_path,
                    summary_path=summary_path,
                )

            with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
                self._write_run_info(writer, metrics)
                self._write_pipeline_summary(writer, metrics)
                if self.config.report_mode in {"standard", "detailed"}:
                    self._write_analysis(writer, analysis_results)
                    self._write_stage_metrics(writer, metrics.stage_metrics)

                if cleaned_df is not None:
                    if self._should_embed_cleaned_data(cleaned_df):
                        cleaned_df.to_excel(writer, sheet_name="Cleaned Data", index=False)
                        if self.config.enable_excel_formatting:
                            self._format_sheet(writer, "Cleaned Data", cleaned_df)
                    else:
                        exported_cleaned_data_path = self._export_cleaned_data(cleaned_df, metrics.run_id)
                        self._write_export_notice(writer, exported_cleaned_data_path)
                elif cleaned_data_path is not None:
                    self._write_export_notice(writer, cleaned_data_path)

                if self.config.enable_excel_formatting:
                    self._format_summary_sheet(writer, "Pipeline Summary")

            LOGGER.info("Report generated.", extra={"stage": "reporting", "report_path": str(report_path)})
            return ReportArtifact(
                report_path=report_path,
                cleaned_data_path=exported_cleaned_data_path,
                drop_reasons_path=drop_reasons_path,
                summary_path=summary_path,
            )
        except Exception as exc:
            raise ReportingError(
                "Report generation failed.",
                error_code="REPORT_GENERATION_FAILED",
                context={"report_path": str(report_path), "run_id": metrics.run_id},
            ) from exc

    def _write_run_info(self, writer: pd.ExcelWriter, metrics: PipelineMetrics) -> None:
        data = {
            "Key": ["Run ID", "Client", "Version", "Created At (UTC)", "Severity"],
            "Value": [metrics.run_id, metrics.client_name, metrics.version, metrics.created_at_utc, metrics.severity],
        }
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name="Run Info", index=False)
        if self.config.enable_excel_formatting:
            self._format_sheet(writer, "Run Info", df)

    def _write_pipeline_summary(self, writer: pd.ExcelWriter, metrics: PipelineMetrics) -> None:
        df = self._build_pipeline_summary_frame(metrics)
        df.to_excel(writer, sheet_name="Pipeline Summary", index=False)
        if self.config.enable_excel_formatting:
            self._format_sheet(writer, "Pipeline Summary", df)

    def _build_pipeline_summary_frame(self, metrics: PipelineMetrics) -> pd.DataFrame:
        invalid_total = (
            metrics.stats.invalid_emails_dropped
            + metrics.stats.invalid_phones_dropped
            + metrics.stats.invalid_numbers_dropped
            + metrics.stats.invalid_dates_dropped
        )
        summary = {
            "Metric": [
                "Rows Loaded",
                "Rows Cleaned",
                "Rows Dropped",
                "Duplicates Dropped",
                "Invalid Rows Dropped (Total)",
                "Drop Rate",
                "Invalid Emails Dropped",
                "Invalid Phones Dropped",
                "Invalid Numbers Dropped",
                "Invalid Dates Dropped",
                "Drop Rate Alert",
                "Email Alert",
                "Phone Alert",
                "Schema Version",
            ],
            "Value": [
                metrics.stats.original_rows,
                metrics.stats.final_rows,
                metrics.dropped_rows,
                metrics.stats.duplicates_dropped,
                invalid_total,
                metrics.drop_rate_pct,
                metrics.stats.invalid_emails_dropped,
                metrics.stats.invalid_phones_dropped,
                metrics.stats.invalid_numbers_dropped,
                metrics.stats.invalid_dates_dropped,
                "EXCEEDED" if metrics.is_drop_rate_alert else "OK",
                "EXCEEDED" if metrics.is_email_alert else "OK",
                "EXCEEDED" if metrics.is_phone_alert else "OK",
                metrics.stats.schema_version,
            ],
        }
        return pd.DataFrame(summary)

    def _write_analysis(self, writer: pd.ExcelWriter, analysis_results: dict[str, Any]) -> None:
        if not analysis_results:
            notice = pd.DataFrame({"message": ["No analysis data available."]})
            notice.to_excel(writer, sheet_name="Analysis", index=False)
            if self.config.enable_excel_formatting:
                self._format_sheet(writer, "Analysis", notice)
            return

        for section, data in analysis_results.items():
            if isinstance(data, dict):
                frame = pd.DataFrame.from_dict(data, orient="index")
                sheet_name = f"A_{section[:28]}"
                frame.to_excel(writer, sheet_name=sheet_name)
                if self.config.enable_excel_formatting:
                    self._format_sheet(writer, sheet_name, frame.reset_index(drop=False))

    def _write_stage_metrics(self, writer: pd.ExcelWriter, stage_metrics: list[dict[str, Any]]) -> None:
        frame = pd.DataFrame(stage_metrics) if stage_metrics else pd.DataFrame({"stage": [], "duration_ms": []})
        frame.to_excel(writer, sheet_name="Stage Metrics", index=False)
        if self.config.enable_excel_formatting:
            self._format_sheet(writer, "Stage Metrics", frame)

    def _write_export_notice(self, writer: pd.ExcelWriter, cleaned_data_path: Path) -> None:
        notice = pd.DataFrame(
            {
                "message": [
                    "Cleaned dataset exported as CSV due to configured row limit.",
                    str(cleaned_data_path),
                ]
            }
        )
        notice.to_excel(writer, sheet_name="Cleaned Data", index=False)
        if self.config.enable_excel_formatting:
            self._format_sheet(writer, "Cleaned Data", notice)

    def _resolve_report_path(self, run_id: str) -> Path:
        suffix_map = {"excel": ".xlsx", "csv": ".csv", "parquet": ".parquet"}
        suffix = suffix_map.get(self.config.report_format, ".xlsx")
        return self.config.paths.reports_dir / f"analysis_report_{run_id}{suffix}"

    def _should_embed_cleaned_data(self, cleaned_df: pd.DataFrame) -> bool:
        return self.config.include_cleaned_data_in_report and len(cleaned_df) <= self.config.max_report_rows

    def _export_cleaned_data(self, cleaned_df: pd.DataFrame, run_id: str) -> Path:
        export_path = self.config.paths.exports_dir / f"cleaned_{run_id}.csv"
        cleaned_df.to_csv(export_path, index=False)
        return export_path

    def _format_summary_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for row in worksheet.iter_rows(min_row=2, max_col=2):
            value_cell = row[1]
            if value_cell.value == "EXCEEDED":
                value_cell.fill = red_fill
            elif value_cell.value == "OK":
                value_cell.fill = green_fill
        _ = workbook

    def _format_sheet(self, writer: pd.ExcelWriter, sheet_name: str, dataframe: pd.DataFrame) -> None:
        worksheet = writer.sheets[sheet_name]
        header_font = Font(bold=True)

        for cell in worksheet[1]:
            cell.font = header_font

        sample = dataframe.head(DEFAULT_COLUMN_WIDTH_SAMPLE_ROWS).fillna("")
        for idx, column in enumerate(sample.columns, start=1):
            col_values = sample[column].astype(str).tolist()
            max_length = max([len(str(column)), *(len(value) for value in col_values)] or [10])
            worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = min(
                max_length + 2,
                DEFAULT_MAX_COLUMN_WIDTH,
            )

    def _resolve_drop_reasons_path(self, run_id: str) -> Path:
        return self.config.paths.reports_dir / f"drop_reasons_{run_id}.csv"

    def _resolve_summary_path(self, run_id: str) -> Path:
        return self.config.paths.reports_dir / f"pipeline_summary_{run_id}.json"

    def _write_drop_reasons(self, metrics: PipelineMetrics, path: Path) -> None:
        invalid_total = (
            metrics.stats.invalid_emails_dropped
            + metrics.stats.invalid_phones_dropped
            + metrics.stats.invalid_numbers_dropped
            + metrics.stats.invalid_dates_dropped
        )
        dropped = metrics.dropped_rows
        rows = metrics.stats.original_rows

        rows_data = [
            ("duplicate", metrics.stats.duplicates_dropped),
            ("invalid_email", metrics.stats.invalid_emails_dropped),
            ("invalid_phone", metrics.stats.invalid_phones_dropped),
            ("invalid_number", metrics.stats.invalid_numbers_dropped),
            ("invalid_date", metrics.stats.invalid_dates_dropped),
            ("invalid_total", invalid_total),
        ]

        data = []
        for reason, count in rows_data:
            data.append(
                {
                    "reason": reason,
                    "count": int(count),
                    "share_of_dropped": (count / dropped) if dropped else 0.0,
                    "share_of_rows": (count / rows) if rows else 0.0,
                }
            )

        pd.DataFrame(data).to_csv(path, index=False)

    def _write_structured_summary(
        self,
        *,
        metrics: PipelineMetrics,
        analysis_results: dict[str, Any],
        report_path: Path,
        cleaned_data_path: Path | None,
        drop_reasons_path: Path,
        summary_path: Path,
    ) -> None:
        meta = analysis_results.get("meta", {}) if analysis_results else {}
        summary = {
            "run_id": metrics.run_id,
            "client_name": metrics.client_name,
            "version": metrics.version,
            "created_at_utc": metrics.created_at_utc,
            "severity": metrics.severity,
            "analysis_mode": meta.get("analysis_mode"),
            "rows_loaded": metrics.stats.original_rows,
            "rows_cleaned": metrics.stats.final_rows,
            "rows_dropped": metrics.dropped_rows,
            "drop_rate": metrics.stats.drop_rate,
            "alerts": {
                "drop_rate": metrics.is_drop_rate_alert,
                "invalid_emails": metrics.is_email_alert,
                "invalid_phones": metrics.is_phone_alert,
            },
            "validation_counts": {
                "duplicates": metrics.stats.duplicates_dropped,
                "invalid_emails": metrics.stats.invalid_emails_dropped,
                "invalid_phones": metrics.stats.invalid_phones_dropped,
                "invalid_numbers": metrics.stats.invalid_numbers_dropped,
                "invalid_dates": metrics.stats.invalid_dates_dropped,
            },
            "schema_version": metrics.stats.schema_version,
            "missing_required_columns": list(metrics.stats.missing_required_columns),
            "extra_columns": list(metrics.stats.extra_columns),
            "paths": {
                "report": str(report_path),
                "cleaned_data": str(cleaned_data_path) if cleaned_data_path else None,
                "drop_reasons": str(drop_reasons_path),
                "summary": str(summary_path),
            },
        }

        summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=True), encoding="utf-8")
