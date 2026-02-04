import os
import logging
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from src.exceptions import ReportError

logger = logging.getLogger(__name__)

REPORT_VERSION = "2.0"

def write_analysis_to_excel(
    results: dict,
    filename: str,
    *,
    clean_df: pd.DataFrame = None,
    cleaning_report: dict = None,
    auto_width_limit: int = 50000
) -> dict:
    """
    Write analysis results (and optionally cleaned data) into an Excel file
    with formatting, metadata, and a pipeline summary sheet.

    Args:
        results: dict containing analysis outputs (meta, statistics, correlation).
        filename: base filename for the report.
        clean_df: cleaned dataframe to include in Cleaned Data sheet.
        cleaning_report: dict with cleaning stats (rows dropped, invalid counts).
        auto_width_limit: max rows to scan for auto column width.

    Returns:
        dict with standardized report info.
    """

    reports_dir = os.getenv("REPORTS_DIR", "data/reports")
    os.makedirs(reports_dir, exist_ok=True)

    # ✅ Timestamped filename to avoid overwrites
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = filename.replace(".xlsx", "")
    filename = f"{base}_{timestamp}.xlsx"
    filepath = Path(reports_dir) / filename

    if not results:
        logger.warning("No analysis results provided.")
        raise ReportError("No analysis results to write.")

    try:
        # ✅ Fixed sheet order (now includes Cleaned Data + Pipeline Summary)
        sheet_order = ["Metadata", "Meta", "Statistics", "Correlation", "Cleaned Data", "Pipeline Summary"]
        sheets = {
            "Meta": results.get("meta"),
            "Statistics": results.get("statistics"),
            "Correlation": results.get("correlation"),
            "Cleaned Data": clean_df,
        }

        # ✅ Write sheets
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Metadata sheet
            metadata = {
                "report_version": REPORT_VERSION,
                "analysis_version": results.get("meta", {}).get("analysis_version", "unknown"),
                "timestamp": timestamp,
                "client_name": os.getenv("CLIENT_NAME", "Unknown"),
                "recipient_email": os.getenv("RECIPIENT_EMAIL", "Unknown"),
            }
            pd.DataFrame.from_dict(metadata, orient="index", columns=["Value"]).to_excel(writer, sheet_name="Metadata")

            # Other sheets
            for sheet_name in sheet_order[1:4]:  # Meta, Statistics, Correlation
                content = sheets.get(sheet_name)
                if content is None or (isinstance(content, dict) and not content):
                    continue
                if isinstance(content, dict):
                    pd.DataFrame.from_dict(content, orient="index").to_excel(writer, sheet_name=sheet_name)
                elif isinstance(content, pd.DataFrame):
                    if not content.empty:
                        content.to_excel(writer, sheet_name=sheet_name, index=False)

            # Cleaned Data sheet
            if clean_df is not None and not clean_df.empty:
                clean_df.to_excel(writer, sheet_name="Cleaned Data", index=False)

            # Pipeline Summary sheet
            if cleaning_report:
                # ✅ Compute drop rate and severity classification (fallback if not already set)
                try:
                    drop_rate = 1 - (cleaning_report["final_rows"] / cleaning_report["original_rows"])
                    if "DROP_RATE" not in cleaning_report:
                        cleaning_report["DROP_RATE"] = f"{drop_rate:.2%}"
                    if "SEVERITY" not in cleaning_report:
                        if drop_rate < 0.10:
                            severity = "LOW"
                        elif drop_rate <= 0.30:
                            severity = "MEDIUM"
                        else:
                            severity = "HIGH ⚠️"
                        cleaning_report["SEVERITY"] = severity
                except Exception as e:
                    logger.warning("Failed to compute drop rate/severity: %s", e)

                # ✅ Build summary dict with thresholds + alerts
                summary_data = {
                    "Rows Loaded": cleaning_report.get("ROWS_LOADED"),
                    "Rows Cleaned": cleaning_report.get("ROWS_CLEANED"),
                    "Drop Rate": cleaning_report.get("DROP_RATE"),
                    "Severity": cleaning_report.get("SEVERITY"),
                    "Drop Rate Threshold": cleaning_report.get("DROP_RATE_THRESHOLD"),
                    "Drop Rate Alert": cleaning_report.get("DROP_RATE_ALERT"),
                    "Invalid Emails Dropped": cleaning_report.get("invalid_emails_dropped", 0),
                    "Invalid Emails Threshold": cleaning_report.get("INVALID_EMAILS_THRESHOLD"),
                    "Invalid Emails Alert": cleaning_report.get("INVALID_EMAILS_ALERT"),
                    "Pipeline Version": cleaning_report.get("PIPELINE_VERSION"),
                }

                pd.DataFrame.from_dict(summary_data, orient="index", columns=["Value"]).to_excel(
                    writer, sheet_name="Pipeline Summary"
                )

        # ✅ Formatting with openpyxl
        wb = load_workbook(filepath)
        for sheet_name in sheet_order:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            # Bold headers
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Auto-width (cap for large datasets)
            row_limit = min(ws.max_row, auto_width_limit)
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col[:row_limit]:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

            # Highlight missing values (skip Cleaned Data + Pipeline Summary)
            if sheet_name.lower() not in ["cleaned data", "pipeline summary"]:
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if cell.value is None or (isinstance(cell.value, float) and pd.isna(cell.value)):
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                        # Number formatting for floats
                        if isinstance(cell.value, float):
                            cell.number_format = "#,##0.00"

        # ✅ Conditional formatting for alerts in Pipeline Summary
        if "Pipeline Summary" in wb.sheetnames:
            ws = wb["Pipeline Summary"]
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                for cell in row:
                    if cell.value == "EXCEEDED":
                        cell.fill = red_fill
                    elif cell.value == "OK":
                        cell.fill = green_fill

        wb.save(filepath)

        logger.info("Excel report written successfully: %s", filepath)

        return {
            "filepath": str(filepath),
            "sheets_written": [s for s in sheet_order if s in wb.sheetnames],
            "timestamp": timestamp,
        }

    except PermissionError as e:
        logger.error("Permission denied: Excel file may be open.")
        raise ReportError("Permission denied: close the Excel file before writing.") from e
    except Exception as e:
        logger.exception("Failed to write Excel report.")
        raise ReportError("Excel writing failed") from e